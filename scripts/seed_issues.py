#!/usr/bin/env python3
"""
Seed issues from Issue_Mapping_Classified.xlsx into Supabase.
Run AFTER running db/migrations/001_create_issues.sql in the Supabase SQL editor.

  python3 scripts/seed_issues.py
"""

import os, sys, json, requests
from pathlib import Path
from datetime import datetime

# ── Load .env.local ────────────────────────────────────────────────────────────
env_file = Path(__file__).parent.parent / '.env.local'
if env_file.exists():
    for line in env_file.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith('#') and '=' in line:
            k, v = line.split('=', 1)
            os.environ.setdefault(k.strip(), v.strip())

SUPABASE_URL = os.environ.get('SUPABASE_URL', '').rstrip('/')
SERVICE_KEY  = os.environ.get('SUPABASE_SERVICE_ROLE_KEY', '')

if not SUPABASE_URL or not SERVICE_KEY:
    print('ERROR: Set SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY in .env.local or env')
    sys.exit(1)

HEADERS = {
    'apikey': SERVICE_KEY,
    'Authorization': f'Bearer {SERVICE_KEY}',
    'Content-Type': 'application/json',
}

# ── Excel path ─────────────────────────────────────────────────────────────────
EXCEL_CANDIDATES = [
    Path('/root/.claude/uploads/0c81ebd0-e529-5880-b904-73683c73dd7d/b7d5ad2c-Issue_Mapping_Classified.xlsx'),
    Path(__file__).parent.parent / 'Issue_Mapping_Classified.xlsx',
]
EXCEL_PATH = next((p for p in EXCEL_CANDIDATES if p.exists()), None)
if not EXCEL_PATH:
    print('ERROR: Cannot find Issue_Mapping_Classified.xlsx')
    sys.exit(1)

# ── Normalisation helpers ──────────────────────────────────────────────────────
def norm_status(s):
    if not s: return 'Open'
    s = str(s).strip()
    sl = s.lower()
    if 'resolv' in sl:    return 'Resolved'
    if 'in progress' in sl or sl == 'in_progress': return 'In Progress'
    if sl == 'closed':    return 'Closed'
    if sl == 'deferred':  return 'Deferred'
    if sl == 'open':      return 'Open'
    return s.strip()

def norm_priority(p):
    if not p: return None
    p = str(p).strip()
    if p in ('P0', 'P1', 'P2', 'P3'): return p
    pl = p.lower()
    if pl in ('high', 'hingh'):  return 'P1'
    if pl == 'medium':           return 'P2'
    if p == 'P)':                return 'P0'
    if not p or p == '\xa0':    return None
    return p

def parse_date(d):
    if not d: return None
    if isinstance(d, datetime): return d.date().isoformat()
    d = str(d).strip()
    if not d: return None
    for fmt in ('%d/%m/%Y', '%d/%m/%y', '%m/%d/%Y', '%Y-%m-%d'):
        try: return datetime.strptime(d, fmt).date().isoformat()
        except: pass
    # "8th April 2026" style
    import re
    m = re.match(r'(\d{1,2})\w*\s+(\w+)\s+(\d{4})', d)
    if m:
        try:
            return datetime.strptime(f"{m.group(1)} {m.group(2)} {m.group(3)}", '%d %B %Y').date().isoformat()
        except: pass
    return None

# ── Fetch accounts for name→id mapping ────────────────────────────────────────
print('Fetching accounts...')
r = requests.get(f'{SUPABASE_URL}/rest/v1/accounts?select=id,account_name&limit=2000', headers=HEADERS)
r.raise_for_status()
accounts = r.json()
name_to_id = {a['account_name'].lower().strip(): a['id'] for a in accounts if a.get('account_name')}
print(f'  {len(accounts)} accounts loaded, {len(name_to_id)} indexed')

# ── Check table exists and is empty ───────────────────────────────────────────
print('Checking issues table...')
check = requests.get(f'{SUPABASE_URL}/rest/v1/issues?select=id&limit=1', headers=HEADERS)
if check.status_code == 404 or 'relation "issues" does not exist' in check.text:
    print('ERROR: issues table does not exist.')
    print('Run db/migrations/001_create_issues.sql in the Supabase SQL editor first.')
    sys.exit(1)
if check.status_code != 200:
    print(f'ERROR: {check.status_code} {check.text[:300]}')
    sys.exit(1)
if check.json():
    count_r = requests.get(f'{SUPABASE_URL}/rest/v1/issues?select=id', headers={**HEADERS, 'Prefer': 'count=exact', 'Range-Unit': 'items', 'Range': '0-0'})
    total = int(count_r.headers.get('Content-Range', '0/0').split('/')[-1])
    print(f'WARNING: issues table already has {total} rows. Skipping to avoid duplicates.')
    print('DELETE FROM issues;  — run this in SQL editor first if you want to re-seed.')
    sys.exit(0)

# ── Read and build records ─────────────────────────────────────────────────────
print(f'Reading {EXCEL_PATH.name}...')
try:
    from openpyxl import load_workbook
except ImportError:
    os.system('pip install openpyxl -q')
    from openpyxl import load_workbook

wb   = load_workbook(EXCEL_PATH, read_only=True)
ws   = wb['Issue Mapping']
rows = list(ws.iter_rows(values_only=True))
hdrs = rows[0]
data = [dict(zip(hdrs, r)) for r in rows[1:] if any(r)]
print(f'  {len(data)} rows in sheet')

records = []
skipped = 0
unmatched_accounts = set()

for row in data:
    desc = (str(row.get('Description') or '')).strip()
    if not desc:
        skipped += 1
        continue

    cname = (str(row.get('Customer Name') or '')).strip()
    account_id = name_to_id.get(cname.lower())
    if cname and not account_id:
        unmatched_accounts.add(cname)

    st = row.get('Support Ticket')
    dt = row.get('Dev Ticket')

    records.append({
        'account_id':    account_id,
        'account_name':  cname or None,
        'tenant_id':     (str(row.get('Tenant Id') or '')).strip() or None,
        'csm_lead':      (str(row.get('CSM Lead') or '')).strip() or None,
        'csm':           (str(row.get('CSM') or '')).strip() or None,
        'description':   desc,
        'priority':      norm_priority(row.get('Priority')),
        'owner_team':    (str(row.get('Owner Team') or '')).strip() or None,
        'support_ticket': int(st) if isinstance(st, (int, float)) and st == int(st) else None,
        'dev_ticket':    int(dt) if isinstance(dt, (int, float)) and dt == int(dt) else None,
        'issue_type':    (str(row.get('Issue Type') or '')).strip() or None,
        'issue_sub_type':(str(row.get('Issue Subtype') or '')).strip() or None,
        'reported_date': parse_date(row.get('Reported Date')),
        'closure_date':  (str(row.get('Closure Date') or '')).strip() or None,
        'status':        norm_status(row.get('Status')),
        'next_steps':    (str(row.get('Next Steps') or '')).strip() or None,
    })

print(f'  {len(records)} records to insert, {skipped} skipped (blank description)')
if unmatched_accounts:
    print(f'  {len(unmatched_accounts)} account names not matched (account_id will be null):')
    for n in sorted(unmatched_accounts)[:10]:
        print(f'    - {n}')
    if len(unmatched_accounts) > 10:
        print(f'    ... and {len(unmatched_accounts)-10} more')

# ── Batch insert ───────────────────────────────────────────────────────────────
BATCH = 100
inserted = 0
print(f'Inserting in batches of {BATCH}...')

for i in range(0, len(records), BATCH):
    batch = records[i:i+BATCH]
    r = requests.post(
        f'{SUPABASE_URL}/rest/v1/issues',
        headers={**HEADERS, 'Prefer': 'return=minimal'},
        data=json.dumps(batch),
    )
    if r.status_code not in (200, 201):
        print(f'\nERROR at batch starting {i}: HTTP {r.status_code}')
        print(r.text[:500])
        sys.exit(1)
    inserted += len(batch)
    pct = int(inserted / len(records) * 100)
    print(f'  {inserted}/{len(records)} ({pct}%)', end='\r', flush=True)

print(f'\nDone! {inserted} issues inserted successfully.')
